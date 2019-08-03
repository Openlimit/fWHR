from fWHR_Caculator import get_fwhr
from os import listdir
import re
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import spider


def calculate(w_sheet, r_sheet, r, image_map):
    for index in range(3):
        w_sheet.cell(row=r + 1, column=index + 1).value = r_sheet.cell(row=r + 1, column=index + 1).value

    compony = r_sheet.cell(row=r + 1, column=2).value
    name = r_sheet.cell(row=r + 1, column=3).value
    # ratio = r_sheet.cell(row=r + 1, column=10).value

    if not compony or not name:
        return

    print(r)
    key = compony + '_' + name
    if key in image_map:
        infos = image_map[key].split('_')
        for index in range(6):
            w_sheet.cell(row=r + 1, column=index + 4).value = infos[index + 2]
        try:
            ratio = get_fwhr('image/' + image_map[key], show=False)
            if ratio:
                w_sheet.cell(row=r + 1, column=10).value = ratio
        except BaseException as exct:
            print(exct)
            print(image_map[key])
        return

    result = {}
    imagePath = ''
    try:
        result = spider.search(compony, name)
        if not result:
            return
        w_sheet.cell(row=r + 1, column=4).value = result['SCO_NAME']
        w_sheet.cell(row=r + 1, column=5).value = result['CER_NUM']
        w_sheet.cell(row=r + 1, column=6).value = result['PTI_NAME']
        w_sheet.cell(row=r + 1, column=7).value = result['ECO_NAME']
        w_sheet.cell(row=r + 1, column=8).value = result['PPP_GET_DATE']
        w_sheet.cell(row=r + 1, column=9).value = result['PPP_END_DATE']

        personID = spider.getPersonID(result['PPP_ID'])
        if not personID:
            return
        imagePath = spider.getImagePath(personID)
        if not imagePath:
            return
    except BaseException as e:
        print(e)
        if str(e) == 'string indices must be integers':
            print("sleep")
            return

    if not result or not imagePath:
        return

    imagename = compony + '_' + name + '_' + result['SCO_NAME'] + '_' + result['CER_NUM'] \
                + '_' + result['PTI_NAME'] + '_' + result['ECO_NAME'] + '_' + \
                result['PPP_GET_DATE'] + '_' + result['PPP_END_DATE']
    try:
        ratio = get_fwhr(imagePath, url=True, show=False, imagename=imagename)
        if ratio:
            w_sheet.cell(row=r + 1, column=10).value = ratio
            print('ok')
    except BaseException as exct:
        print(exct)
        print(result)
        print(imagePath)


def calculate_simple(w_sheet, r_sheet, r):
    print(r)
    for index in range(3):
        w_sheet.cell(row=r + 1, column=index + 1).value = r_sheet.cell(row=r + 1, column=index + 1).value

    compony = r_sheet.cell(row=r + 1, column=3).value
    name = r_sheet.cell(row=r + 1, column=2).value
    ratio = r_sheet.cell(row=r + 1, column=6).value

    if ratio:
        print('pass')
        return
    if not compony or not name:
        return

    result = {}
    imagePath = ''
    try:
        result = spider.search(compony, name)
        if not result:
            return
        w_sheet.cell(row=r + 1, column=4).value = result['SCO_NAME']
        w_sheet.cell(row=r + 1, column=5).value = result['ECO_NAME']

        personID = result['personID']
        if not personID:
            return
        imagePath = spider.getImagePath(personID)
        if not imagePath:
            return
    except BaseException as e:
        print(e)
        if str(e) == 'string indices must be integers':
            print("sleep")
            return

    if not result or not imagePath:
        return

    imagename = compony + '_' + name + '_' + result['SCO_NAME'] + '_' + result['CER_NUM'] \
                + '_' + result['PTI_NAME'] + '_' + result['ECO_NAME'] + '_' + \
                result['PPP_GET_DATE'] + '_' + result['PPP_END_DATE']
    try:
        ratio = get_fwhr(imagePath, url=True, show=False, imagename=imagename)
        if ratio:
            w_sheet.cell(row=r + 1, column=6).value = ratio
            print('ok: {}'.format(ratio))
    except BaseException as exct:
        print(exct)
        print(result)
        print(imagePath)


def merge_excel(filename1, filename2):
    r_book1 = openpyxl.load_workbook(filename1)
    r_sheet1 = r_book1['Sheet1']
    rows = r_sheet1.max_row

    r_book2 = openpyxl.load_workbook(filename2)
    r_sheet2 = r_book2['Sheet1']

    pattern = re.compile(r'(.*?_.*?)_.*')
    filepath = '/home/wly/python_item/fWHR/box'
    imagename_list = listdir(filepath)
    image_map = {}

    for imagename in imagename_list:
        if imagename.endswith('gen.jpg'):
            continue
        result = pattern.match(imagename)
        if result:
            name = result.group(1)
            image_map[name] = imagename

    w_book = openpyxl.Workbook()
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'brokern'
    w_sheet.cell(row=1, column=2).value = 'brokercd'
    w_sheet.cell(row=1, column=3).value = 'ananm'
    w_sheet.cell(row=1, column=4).value = '性别'
    w_sheet.cell(row=1, column=5).value = '证书编号'
    w_sheet.cell(row=1, column=6).value = '执业岗位'
    w_sheet.cell(row=1, column=7).value = '学历'
    w_sheet.cell(row=1, column=8).value = '证书取得日期'
    w_sheet.cell(row=1, column=9).value = '证书有效截止日期'
    w_sheet.cell(row=1, column=10).value = 'ratio'
    w_sheet.cell(row=1, column=11).value = 'is_cm'

    for r in range(rows):
        if r == 0:
            continue
        for index in range(11):
            w_sheet.cell(row=r + 1, column=index + 1).value = r_sheet1.cell(row=r + 1, column=index + 1).value

        ratio1 = r_sheet1.cell(row=r + 1, column=10).value
        ratio2 = r_sheet2.cell(row=r + 1, column=10).value

        key = r_sheet1.cell(row=r + 1, column=1).value + '_' + r_sheet1.cell(row=r + 1, column=3).value

        if not ratio1 or not ratio2:
            continue
        if key in image_map:
            w_sheet.cell(row=r + 1, column=10).value = ratio2

    w_book.save('/home/wly/python_item/sample_out_up_merge.xlsx')


def upate_excel(filename):
    pattern = re.compile(r'(.*?_.*?)_.*')
    filepath = 'image'

    imagename_list = listdir(filepath)
    image_map = {}
    for imagename in imagename_list:
        if imagename.endswith('gen.jpg'):
            continue
        result = pattern.match(imagename)
        if result:
            name = result.group(1)
            image_map[name] = imagename

    r_book = openpyxl.load_workbook(filename)
    r_sheet = r_book['Sheet1']
    rows = r_sheet.max_row

    w_book = openpyxl.Workbook()  # 创建excel对象
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'brokern'
    w_sheet.cell(row=1, column=2).value = 'brokercd'
    w_sheet.cell(row=1, column=3).value = 'ananm'
    w_sheet.cell(row=1, column=4).value = '性别'
    w_sheet.cell(row=1, column=5).value = '证书编号'
    w_sheet.cell(row=1, column=6).value = '执业岗位'
    w_sheet.cell(row=1, column=7).value = '学历'
    w_sheet.cell(row=1, column=8).value = '证书取得日期'
    w_sheet.cell(row=1, column=9).value = '证书有效截止日期'
    w_sheet.cell(row=1, column=10).value = 'ratio'

    with ThreadPoolExecutor(8) as executor:
        for r in range(rows):
            if r == 0:
                continue
            executor.submit(calculate, w_sheet, r_sheet, r, image_map)

    w_book.save(filename[:-5] + '_out.xlsx')


def udpate_simple(filename):
    r_book = openpyxl.load_workbook(filename)
    r_sheet = r_book['Sheet1']
    rows = r_sheet.max_row

    w_book = openpyxl.Workbook()  # 创建excel对象
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'brokercd'
    w_sheet.cell(row=1, column=2).value = 'ananm'
    w_sheet.cell(row=1, column=3).value = 'brokern'
    w_sheet.cell(row=1, column=4).value = '性别'
    w_sheet.cell(row=1, column=5).value = '学历'
    w_sheet.cell(row=1, column=6).value = 'var12'

    with ThreadPoolExecutor(8) as executor:
        for r in range(rows):
            if r == 0:
                continue
            executor.submit(calculate_simple, w_sheet, r_sheet, r)

    w_book.save(filename[:-5] + '_out.xlsx')


if __name__ == '__main__':
    udpate_simple('wang_out.xlsx')
