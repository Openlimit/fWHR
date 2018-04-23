import requests
import json
from PIL import Image, ImageDraw
import face_recognition
import math
from matplotlib.pyplot import imshow
import pylab
from os import listdir
import re
import openpyxl
from concurrent.futures import ThreadPoolExecutor

request_param = {
    'api_key': '5TzQcAuH_FpsdPuxnnArMvv3FjwtSkoZ',
    'api_secret': 'jHoJXzgoTopkI_GOGJvabF46WCDOP0Go',
    'return_landmark': 1
}

detect_url = 'https://api-cn.faceplusplus.com/facepp/v3/detect'


def detect(filepath):
    files = {
        'image_file': ('test', open(filepath, 'rb'), 'image/jpeg')
    }
    result = requests.post(detect_url, request_param, files=files)

    if not result.text:
        return None
    res = json.loads(result.text)

    if not res['faces']:
        return None

    return res['faces'][0]['landmark']


def cal_fWHR(filepath):
    landmark = detect(filepath)
    if not landmark:
        return None

    width_left, width_right = landmark['contour_left1'], landmark['contour_right1']
    top_left, top_right = landmark['left_eye_top'], landmark['right_eye_top']
    bottom = landmark['mouth_upper_lip_top']

    top_average = (top_left['y'] + top_right['y']) / 2.0

    width = width_right['x'] - width_left['x']
    height = bottom['y'] - top_average
    ratio = float(width) / float(height)

    # corners = {
    #     'top_left': (width_left['x'], top_average),
    #     'bottom_left': (width_left['x'], bottom['y']),
    #     'top_right': (width_right['x'], top_average),
    #     'bottom_right': (width_right['x'], bottom['y'])
    # }
    # show_box(corners)

    return ratio


def show_box(corners):
    image = face_recognition.load_image_file('test')
    pil_image = Image.fromarray(image)
    w, h = pil_image.size

    ## Automatically determine width of the line depending on size of picture
    line_width = math.ceil(h / 200)

    d = ImageDraw.Draw(pil_image)
    d.line([corners['bottom_left'], corners['top_left']], width=line_width)
    d.line([corners['bottom_left'], corners['bottom_right']], width=line_width)
    d.line([corners['top_left'], corners['top_right']], width=line_width)
    d.line([corners['top_right'], corners['bottom_right']], width=line_width)

    imshow(pil_image)
    pylab.show()


def cal_one(w_sheet, r_sheet, r, image_map):
    for index in range(3):
        w_sheet.cell(row=r + 1, column=index + 1).value = r_sheet.cell(row=r + 1, column=index + 1).value

    compony = r_sheet.cell(row=r + 1, column=1).value
    name = r_sheet.cell(row=r + 1, column=3).value
    key = compony + '_' + name

    if key in image_map:
        print(r)
        infos = image_map[key].split('_')
        for index in range(6):
            w_sheet.cell(row=r + 1, column=index + 4).value = infos[index + 2]
        try:
            ratio = cal_fWHR('image/' + image_map[key])
            if ratio:
                w_sheet.cell(row=r + 1, column=10).value = ratio
                print('ok')
        except BaseException as exct:
            print(exct)
            print(image_map[key])


def update_one(w_sheet, r_sheet, r, image_map):
    for index in range(10):
        w_sheet.cell(row=r + 1, column=index + 1).value = r_sheet.cell(row=r + 1, column=index + 1).value

    ratio = r_sheet.cell(row=r + 1, column=10).value
    if not ratio:
        return
    if ratio == 2:
        w_sheet.cell(row=r + 1, column=10).value = 0
        return

    compony = r_sheet.cell(row=r + 1, column=1).value
    name = r_sheet.cell(row=r + 1, column=3).value
    key = compony + '_' + name
    old_ratio = ratio

    if key in image_map:
        print(r)
        try:
            ratio = cal_fWHR('image/' + image_map[key])
            if ratio:
                w_sheet.cell(row=r + 1, column=10).value = ratio
                print(key + ' ' + str(old_ratio) + ' ' + str(ratio))
        except BaseException as exct:
            print(exct)
            print(image_map[key])


def cal_all(filename):
    pattern = re.compile(r'(.*?_.*?)_.*')
    filepath = 'image/'

    imagename_list = listdir(filepath)
    image_map = {}
    for imagename in imagename_list:
        if imagename.endswith('gen.jpg'):
            continue
        result = pattern.match(imagename)
        if result:
            name = result.group(1)
            image_map[name] = imagename

    r_book = openpyxl.load_workbook(filename)
    r_sheet = r_book['Sheet1']
    rows = r_sheet.max_row

    w_book = openpyxl.Workbook()  # 创建excel对象
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'brokern'
    w_sheet.cell(row=1, column=2).value = 'brokercd'
    w_sheet.cell(row=1, column=3).value = 'ananm'
    w_sheet.cell(row=1, column=4).value = '性别'
    w_sheet.cell(row=1, column=5).value = '证书编号'
    w_sheet.cell(row=1, column=6).value = '执业岗位'
    w_sheet.cell(row=1, column=7).value = '学历'
    w_sheet.cell(row=1, column=8).value = '证书取得日期'
    w_sheet.cell(row=1, column=9).value = '证书有效截止日期'
    w_sheet.cell(row=1, column=10).value = 'ratio'

    with ThreadPoolExecutor(16) as executor:
        for r in range(rows):
            if r == 0:
                continue
            executor.submit(cal_one, w_sheet, r_sheet, r, image_map)

    w_book.save('sample_face++.xlsx')


def update(filename):
    pattern = re.compile(r'(.*?_.*?)_.*')
    filepath = 'image/'

    imagename_list = listdir(filepath)
    image_map = {}
    for imagename in imagename_list:
        if imagename.endswith('gen.jpg'):
            continue
        result = pattern.match(imagename)
        if result:
            name = result.group(1)
            image_map[name] = imagename

    r_book = openpyxl.load_workbook(filename)
    r_sheet = r_book['Sheet1']
    rows = r_sheet.max_row

    w_book = openpyxl.Workbook()  # 创建excel对象
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'brokern'
    w_sheet.cell(row=1, column=2).value = 'brokercd'
    w_sheet.cell(row=1, column=3).value = 'ananm'
    w_sheet.cell(row=1, column=4).value = '性别'
    w_sheet.cell(row=1, column=5).value = '证书编号'
    w_sheet.cell(row=1, column=6).value = '执业岗位'
    w_sheet.cell(row=1, column=7).value = '学历'
    w_sheet.cell(row=1, column=8).value = '证书取得日期'
    w_sheet.cell(row=1, column=9).value = '证书有效截止日期'
    w_sheet.cell(row=1, column=10).value = 'ratio'

    with ThreadPoolExecutor(8) as executor:
        for r in range(rows):
            if r == 0:
                continue
            executor.submit(update_one, w_sheet, r_sheet, r, image_map)

    # for r in range(rows):
    #     if r == 0:
    #         continue
    #     update_one(w_sheet, r_sheet, r, image_map)

    w_book.save('sample_face++_update.xlsx')


def statistic(filename):
    r_book = openpyxl.load_workbook(filename)
    r_sheet = r_book['Sheet1']
    rows = r_sheet.max_row

    st_map = {}
    for r in range(rows):
        if r == 0:
            continue
        ratio = r_sheet.cell(row=r + 1, column=2).value
        if not ratio:
            continue
        ratio = round(ratio, 2)
        if ratio in st_map:
            st_map[ratio] += 1
        else:
            st_map[ratio] = 1

    w_book = openpyxl.Workbook()  # 创建excel对象
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'ratio'
    w_sheet.cell(row=1, column=2).value = 'num'
    r = 2
    for ratio in st_map:
        w_sheet.cell(row=r, column=1).value = ratio
        w_sheet.cell(row=r, column=2).value = st_map[ratio]
        r += 1

    w_book.save(filename[:-5] + '_st.xlsx')


def cal_lfw_one(w_sheet, r, imagename, dir_path):
    print(r)
    try:
        w_sheet.cell(row=r, column=1).value = imagename
        ratio = cal_fWHR(dir_path + '/' + imagename)
        if ratio:
            w_sheet.cell(row=r, column=2).value = ratio
            print('ok')
    except BaseException as exct:
        print(exct)


def cal_lfw():
    lfw_path = 'lfw/'
    dir_list = listdir(lfw_path)

    w_book = openpyxl.Workbook()  # 创建excel对象
    w_sheet = w_book.create_sheet(title='Sheet1')
    w_sheet.cell(row=1, column=1).value = 'name'
    w_sheet.cell(row=1, column=2).value = 'ratio'

    r = 2
    with ThreadPoolExecutor(12) as executor:
        for dirname in dir_list:
            image_list = listdir(lfw_path + dirname)
            for imagename in image_list:
                executor.submit(cal_lfw_one, w_sheet, r, imagename, lfw_path + dirname)
                r += 1

    w_book.save('lfw.xlsx')


if __name__ == '__main__':
    statistic('lfw.xlsx')
